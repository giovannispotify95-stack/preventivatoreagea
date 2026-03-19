"""Quick analysis of the None columns in Generali Tariffario."""
import openpyxl

wb = openpyxl.load_workbook(
    "/Users/giovannipucariello/Desktop/PreventivatoreAndre/preventivatoreagea/resources/Generali.xlsx",
    read_only=True, data_only=True
)
ws = wb["Tariffario"]

# Row 1 and 2 for all 38 columns
for r in [1, 2]:
    print(f"Row {r}:")
    for c in range(1, 39):
        val = ws.cell(r, c).value
        if val is not None:
            print(f"  Col {c}: {val}")
    print()

# Row 3 sample data
print("Row 3 (data):")
for c in range(1, 39):
    val = ws.cell(3, c).value
    print(f"  Col {c}: {val}")

wb.close()
