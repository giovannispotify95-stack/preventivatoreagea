"""Script to write all three parser files."""
import os

BASE = "/Users/giovannipucariello/Desktop/PreventivatoreAndre/preventivatoreagea/backend/app/parsers"

# ============== GENERALI ==============
generali = '''\
"""
Parser per Generali.xlsx — sheet "Tariffario".

Layout (anno 2026):
  Riga 1: intestazioni macro-area
  Riga 2: nomi colonna veri
  Riga 3+: dati
"""
from __future__ import annotations

from typing import IO

import openpyxl
from sqlalchemy.orm import Session

from app.models import Tariffa
from app.calcolo import TIPO_GARANZIA

# Mappatura posizionale 1-based -> (franchigia, agevolato, non_agev, totale)
GARANZIE_MAP = {
    "grandine":                (12, 13, 14, 15),
    "eccesso_pioggia":         (None, 16, 17, 18),
    "vento_forte":             (22, 23, 24, 25),
    "eccesso_neve":            (None, 26, 27, 28),
    "gelo_brina":              (29, None, None, 30),
    "siccita":                 (29, None, None, 31),
    "alluvione":               (29, None, None, 32),
    "colpo_sole_vento_caldo":  (None, 33, 34, 35),
    "sbalzo_termico":          (None, 36, 37, 38),
}

FRANCHIGIA_FISSA = {
    "eccesso_pioggia": 30.0,
    "eccesso_neve": 30.0,
    "colpo_sole_vento_caldo": 30.0,
    "sbalzo_termico": 30.0,
}


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def parse_generali(file, db, anno=2026, versione=1):
    """Parsa Generali.xlsx sheet Tariffario. Restituisce record inseriti."""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb["Tariffario"]
    count = 0

    for row_idx in range(3, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        comune_istat_raw = _safe_str(get(4))
        if not comune_istat_raw:
            continue

        comune_istat = comune_istat_raw.zfill(6) if comune_istat_raw.isdigit() else comune_istat_raw
        provincia      = _safe_str(get(2))
        comune_ciag    = _safe_str(get(3))
        comune_nome    = _safe_str(get(5))
        specie_cod     = _safe_str(get(6))
        specie_desc    = _safe_str(get(9))
        raggruppamento = _safe_str(get(11))
        fr_cat = _safe_float(get(29))

        for garanzia, (fr_col, ag_col, na_col, tot_col) in GARANZIE_MAP.items():
            tasso_ag  = _safe_float(get(ag_col))  if ag_col  else 0.0
            tasso_na  = _safe_float(get(na_col))  if na_col  else 0.0
            tasso_tot = _safe_float(get(tot_col)) if tot_col else 0.0

            if abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9 and abs(tasso_tot) < 1e-9:
                continue

            if abs(tasso_tot) < 1e-9:
                tasso_tot = round(tasso_ag + tasso_na, 4)

            if abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                tasso_ag = tasso_tot

            if fr_col is not None:
                franchigia = _safe_float(get(fr_col))
            else:
                franchigia = FRANCHIGIA_FISSA.get(garanzia, fr_cat)

            db.add(Tariffa(
                compagnia="Generali",
                provincia=provincia,
                comune_istat=comune_istat,
                comune_ciag=comune_ciag,
                comune_nome=comune_nome,
                specie_codice=specie_cod,
                specie_descrizione=specie_desc,
                raggruppamento=raggruppamento,
                garanzia=garanzia,
                tipo_garanzia=TIPO_GARANZIA.get(garanzia, "frequenziale"),
                franchigia_min=franchigia,
                franchigia_applicata=franchigia,
                tasso_agevolato=tasso_ag,
                tasso_non_agevolato=tasso_na,
                tasso_totale=tasso_tot,
                anno_validita=anno,
                versione_listino=versione,
            ))
            count += 1

    wb.close()
    db.commit()
    return count
'''

# ============== REVO ==============
revo = '''\
"""
Parser per REVO.xlsx
Sheet "tutti i prodotti": Garanzie Frequenziali per Provincia/Comune/Specie
  Riga 4 = header; Riga 5+ = dati
  Col layout (1-based):
    1=Provincia, 2=Codice Comune, 3=Comune, 4=Raggruppamento,
    5=Codice Specie, 6=Specie,
    7=GRANDINE(label), 8=Fr Min GR, 9=Tasso Ag GR, 10=Tasso NA GR,
    11=VENTO FORTE(label), 12=Fr Min VF, 13=Tasso Ag VF, 14=Tasso NA VF,
    15=ECCESSO DI PIOGGIA(label), 16=Fr Min EP, 17=Tasso Ag EP, 18=Tasso NA EP

Sheet "altre garanzie": Catastrofali + extra per Raggruppamento
  Riga 1 = header; Riga 2+ = dati
  Col: 1=Raggruppamento, 2=GARANZIA, 3=Franchigia Min, 4=Tasso Ag, 5=Tasso NA
"""
from __future__ import annotations

from typing import IO

import openpyxl
from sqlalchemy.orm import Session

from app.models import Tariffa
from app.calcolo import TIPO_GARANZIA, normalizza_garanzia


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


# Blocchi frequenziali: (garanzia_norm, fr_col, ag_col, na_col)
FREQ_BLOCKS = [
    ("grandine",        8,  9, 10),
    ("vento_forte",    12, 13, 14),
    ("eccesso_pioggia",16, 17, 18),
]


def _parse_tutti_i_prodotti(wb, db, anno, versione):
    """Sheet 'tutti i prodotti': frequenziali per Provincia/Comune/Specie."""
    ws = wb["tutti i prodotti"]
    count = 0

    for row_idx in range(5, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        cod_comune = _safe_str(get(2))
        if not cod_comune:
            continue

        provincia      = _safe_str(get(1))
        comune_nome    = _safe_str(get(3))
        raggruppamento = _safe_str(get(4))
        specie_cod     = _safe_str(get(5))
        specie_desc    = _safe_str(get(6))

        for garanzia, fr_col, ag_col, na_col in FREQ_BLOCKS:
            fr = _safe_float(get(fr_col))
            ag = _safe_float(get(ag_col))
            na = _safe_float(get(na_col))

            if abs(ag) < 1e-9 and abs(na) < 1e-9:
                continue

            db.add(Tariffa(
                compagnia="REVO",
                provincia=provincia,
                comune_istat="",
                comune_ciag=cod_comune,
                comune_nome=comune_nome,
                specie_codice=specie_cod,
                specie_descrizione=specie_desc,
                raggruppamento=raggruppamento,
                garanzia=garanzia,
                tipo_garanzia="frequenziale",
                franchigia_min=fr,
                franchigia_applicata=fr,
                tasso_agevolato=ag,
                tasso_non_agevolato=na,
                tasso_totale=round(ag + na, 4),
                anno_validita=anno,
                versione_listino=versione,
            ))
            count += 1

    return count


def _parse_altre_garanzie(wb, db, anno, versione):
    """Sheet 'altre garanzie': catastrofali + extra per raggruppamento."""
    ws = wb["altre garanzie"]
    count = 0

    for row_idx in range(2, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        raggruppamento = _safe_str(get(1))
        garanzia_raw   = _safe_str(get(2))
        if not garanzia_raw:
            continue

        garanzia_norm = normalizza_garanzia(garanzia_raw, "revo")
        if not garanzia_norm or garanzia_norm not in TIPO_GARANZIA:
            # Es. "TAB D EXTRA Q", "cod 005A" -> non mappabili, skip
            continue

        fr = _safe_float(get(3))
        ag = _safe_float(get(4))
        na = _safe_float(get(5))

        if abs(ag) < 1e-9 and abs(na) < 1e-9:
            continue

        db.add(Tariffa(
            compagnia="REVO",
            provincia="",
            comune_istat="",
            comune_ciag="",
            comune_nome="",
            specie_codice="",
            specie_descrizione="",
            raggruppamento=raggruppamento,
            garanzia=garanzia_norm,
            tipo_garanzia=TIPO_GARANZIA.get(garanzia_norm, "catastrofale"),
            franchigia_min=fr,
            franchigia_applicata=fr,
            tasso_agevolato=ag,
            tasso_non_agevolato=na,
            tasso_totale=round(ag + na, 4),
            anno_validita=anno,
            versione_listino=versione,
        ))
        count += 1

    return count


def parse_revo(file, db, anno=2026, versione=1):
    """Parsa entrambi gli sheet di REVO.xlsx. Restituisce record totali."""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    c1 = _parse_tutti_i_prodotti(wb, db, anno, versione)
    c2 = _parse_altre_garanzie(wb, db, anno, versione)
    wb.close()
    db.commit()
    return c1 + c2
'''

# ============== REALE MUTUA ==============
reale_mutua = '''\
"""
Parser per RealeMutua.xlsx — sheet "ARROTONDA".

66 colonne. Layout:
  Col  1  codice ISTAT comune      Col 37  Tasso GB intero
  Col  2  zona AB comune           Col 38  Tasso GB Agevolato
  Col  3  codice CIAG comune       Col 39  Tasso GB Non agev
  Col  4  Descrizione comune       Col 40  Tasso GB_prim_aut intero
  Col  5  codice specie            Col 41  Tasso GB_prim_aut Agevolato
  Col  6  descrizione specie       Col 42  Tasso GB_prim_aut Non agev
  Col  7  Fr Minima GR             Col 43  Tasso SI irrigua intero
  Col  8  Fr Minima VF             Col 44  Tasso SI irrigua Agevolato
  Col  9  Fr Minima CAT            Col 45  Tasso SI irrigua Non agev
  Col 10-12  GR fr10 (int/ag/na)   Col 46  Tasso SI NON irrigua intero
  Col 13-15  GR fr15 (int/ag/na)   Col 47  Tasso SI NON irrigua Agevolato
  Col 16-18  GR fr20 (int/ag/na)   Col 48  Tasso SI NON irrigua Non agev
  Col 19-21  GR fr30 (int/ag/na)   Col 49  Tasso AL intero
  Col 22-24  VF fr10 (int/ag/na)   Col 50  Tasso AL Agevolato
  Col 25-27  VF fr15 (int/ag/na)   Col 51  Tasso AL Non agev
  Col 28-30  VF fr20 (int/ag/na)   Col 52  Tasso ST Non agev fr 30%
  Col 31-33  VF fr30 (int/ag/na)   Col 53  Tasso EN intero fr 30%
  Col 34-36  EP fr30 (int/ag/na)   Col 54  Tasso EN Agevolato fr 30%
                                    Col 55  Tasso EN Non agev fr 30%
  Col 56  Tasso CS-VC intero fr 30%
  Col 57  Tasso CS-VC Agevolato fr 30%
  Col 58  Tasso CS-VC Non agev fr 30%
  Col 59  Tasso Accessorie intero
  Col 60  Tasso Accessorie Agevolato
  Col 61  Tasso Accessorie Non Agev
  Col 62  Cod_Tipologia_Prodotto
  Col 64  Raggruppamento specie

Deduplicazione: per duplicati su (codice ISTAT, codice specie)
si tiene la riga con la somma dei tassi piu bassa.
"""
from __future__ import annotations

from typing import IO

import openpyxl
import pandas as pd
from sqlalchemy.orm import Session

from app.models import Tariffa
from app.calcolo import TIPO_GARANZIA


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


# Mappatura garanzie con multi-franchigia: (garanzia, [(fr%, int_col, ag_col, na_col), ...])
GARANZIE_MULTI_FR = {
    "grandine": [
        (10, 10, 11, 12),
        (15, 13, 14, 15),
        (20, 16, 17, 18),
        (30, 19, 20, 21),
    ],
    "vento_forte": [
        (10, 22, 23, 24),
        (15, 25, 26, 27),
        (20, 28, 29, 30),
        (30, 31, 32, 33),
    ],
    "eccesso_pioggia": [
        (30, 34, 35, 36),
    ],
}

# Garanzie con franchigia unica (catastrofali e accessorie)
GARANZIE_SINGOLE = {
    "gelo_brina":             (37, 38, 39),   # GB intero/ag/na
    "siccita":                (43, 44, 45),   # SI irrigua
    "alluvione":              (49, 50, 51),   # AL
    "eccesso_neve":           (53, 54, 55),   # EN
    "colpo_sole_vento_caldo": (56, 57, 58),   # CS-VC
    "sbalzo_termico":         (None, None, 52),  # ST ha solo non_agev fr 30%
}


def parse_reale_mutua(file, db, anno=2026, versione=1):
    """Parsa RealeMutua.xlsx sheet ARROTONDA. Deduplica e inserisce tariffe."""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb["ARROTONDA"]

    # Leggi tutte le righe in una lista di dict
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        istat_raw = _safe_str(get(1))
        specie_raw = _safe_str(get(5))
        if not istat_raw and not specie_raw:
            continue

        # Calcola somma totale tassi per deduplicazione
        tasso_sum = 0.0
        for c in range(10, 62):
            tasso_sum += abs(_safe_float(get(c)))

        rows.append({
            "row_idx": row_idx,
            "istat": istat_raw,
            "ciag": _safe_str(get(3)),
            "comune_nome": _safe_str(get(4)),
            "specie_cod": specie_raw,
            "specie_desc": _safe_str(get(6)),
            "fr_gr": _safe_float(get(7)),
            "fr_vf": _safe_float(get(8)),
            "fr_cat": _safe_float(get(9)),
            "raggruppamento": _safe_str(get(64)),
            "tasso_sum": tasso_sum,
        })

    wb.close()

    # Deduplicazione: per (ISTAT, codice specie) tieni riga con tassi piu bassi
    best = {}
    for r in rows:
        key = (r["istat"], r["specie_cod"])
        if key not in best or r["tasso_sum"] < best[key]["tasso_sum"]:
            best[key] = r

    # Riapri il workbook per leggere i tassi delle righe selezionate
    file.seek(0)
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb["ARROTONDA"]

    count = 0
    for r in best.values():
        row_idx = r["row_idx"]
        get = lambda c: ws.cell(row_idx, c).value

        istat = r["istat"]
        if istat and istat.replace(".", "").isdigit():
            istat = str(int(float(istat))).zfill(6)

        # Garanzie con multi-franchigia (GR, VF, EP)
        for garanzia, fr_list in GARANZIE_MULTI_FR.items():
            fr_min = r["fr_gr"]
            if garanzia == "vento_forte":
                fr_min = r["fr_vf"]

            for fr_pct, int_col, ag_col, na_col in fr_list:
                tasso_int = _safe_float(get(int_col))
                tasso_ag  = _safe_float(get(ag_col))
                tasso_na  = _safe_float(get(na_col))

                if abs(tasso_int) < 1e-9 and abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                    continue

                if abs(tasso_int) < 1e-9:
                    tasso_int = round(tasso_ag + tasso_na, 4)

                db.add(Tariffa(
                    compagnia="RealeMutua",
                    provincia="",
                    comune_istat=istat,
                    comune_ciag=r["ciag"],
                    comune_nome=r["comune_nome"],
                    specie_codice=r["specie_cod"],
                    specie_descrizione=r["specie_desc"],
                    raggruppamento=r["raggruppamento"],
                    garanzia=garanzia,
                    tipo_garanzia=TIPO_GARANZIA.get(garanzia, "frequenziale"),
                    franchigia_min=fr_min,
                    franchigia_applicata=float(fr_pct),
                    tasso_agevolato=tasso_ag,
                    tasso_non_agevolato=tasso_na,
                    tasso_totale=tasso_int,
                    anno_validita=anno,
                    versione_listino=versione,
                ))
                count += 1

        # Garanzie singole (catastrofali)
        fr_cat = r["fr_cat"]
        for garanzia, (int_col, ag_col, na_col) in GARANZIE_SINGOLE.items():
            tasso_int = _safe_float(get(int_col)) if int_col else 0.0
            tasso_ag  = _safe_float(get(ag_col))  if ag_col  else 0.0
            tasso_na  = _safe_float(get(na_col))  if na_col  else 0.0

            if abs(tasso_int) < 1e-9 and abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                continue

            if abs(tasso_int) < 1e-9:
                tasso_int = round(tasso_ag + tasso_na, 4)
            if abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                tasso_ag = tasso_int

            db.add(Tariffa(
                compagnia="RealeMutua",
                provincia="",
                comune_istat=istat,
                comune_ciag=r["ciag"],
                comune_nome=r["comune_nome"],
                specie_codice=r["specie_cod"],
                specie_descrizione=r["specie_desc"],
                raggruppamento=r["raggruppamento"],
                garanzia=garanzia,
                tipo_garanzia=TIPO_GARANZIA.get(garanzia, "catastrofale"),
                franchigia_min=fr_cat,
                franchigia_applicata=fr_cat if fr_cat > 0 else 30.0,
                tasso_agevolato=tasso_ag,
                tasso_non_agevolato=tasso_na,
                tasso_totale=tasso_int,
                anno_validita=anno,
                versione_listino=versione,
            ))
            count += 1

    wb.close()
    db.commit()
    return count
'''

# Write all files
for name, content in [("generali.py", generali), ("revo.py", revo), ("reale_mutua.py", reale_mutua)]:
    path = os.path.join(BASE, name)
    with open(path, "w") as f:
        f.write(content)
    print(f"Written {path} ({len(content)} chars)")

print("Done!")
