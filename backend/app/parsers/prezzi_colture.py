"""
Parser per Prezzi_colture.xlsx — sheet "Tabelle1".

Layout:
  Riga 1: titolo consorzio
  Riga 2: macro-header  (CODICE PRODOTTO, DESCRIZIONE PRODOTTO, ...)
  Riga 3: sotto-header   (CIAG, ANIA, MIPAAF, ..., CONSORZIO, ISMEA, MAX, MED, MIN, ...)
  Riga 4+: dati

Colonne (1-based):
  1  = CIAG
  2  = ANIA
  3  = MIPAAF
  4  = Descrizione prodotto
  5  = Descrizione varieta
  6  = Codice consorzio
  7  = Codice ISMEA
  8  = Prezzo MAX
  9  = Prezzo MED
  10 = Prezzo MIN
  11 = Territorio
  12 = Standard Value 2026
  13 = Indicazione SV provvisorio
  14 = Coefficiente maggiorazione
  15 = Standard Value BIO 2026
"""
from __future__ import annotations

from typing import IO

import openpyxl
from sqlalchemy.orm import Session

from app.models import PrezzoColtura


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def parse_prezzi_colture(
    file: IO, db: Session, anno: int = 2026
) -> int:
    """Parsa Prezzi_colture.xlsx e inserisce i prezzi nel database."""
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    count = 0

    for row_idx in range(4, ws.max_row + 1):
        get = lambda c, _r=row_idx: ws.cell(_r, c).value

        ciag = _safe_str(get(1))
        if not ciag:
            continue

        desc = _safe_str(get(4))
        if not desc:
            continue

        prezzo = PrezzoColtura(
            codice_ciag=ciag,
            codice_ania=_safe_str(get(2)),
            codice_mipaaf=_safe_str(get(3)),
            descrizione=desc,
            varieta=_safe_str(get(5)),
            prezzo_consorzio=_safe_float(get(6)),
            prezzo_ismea=_safe_float(get(7)),
            prezzo_max=_safe_float(get(8)),
            prezzo_med=_safe_float(get(9)),
            prezzo_min=_safe_float(get(10)),
            coeff_maggiorazione=_safe_float(get(14)) or 1.0,
            standard_value_bio=_safe_float(get(15)),
            anno=anno,
        )
        db.add(prezzo)
        count += 1

    wb.close()
    db.commit()
    return count
