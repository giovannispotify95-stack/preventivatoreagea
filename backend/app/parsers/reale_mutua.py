"""
Parser per RealeMutua.xlsx — sheet "Sconti" e "Tariffa_Normale".

Entrambi gli sheet condividono il medesimo layout a 62 colonne di tassi:
  Col  1  codice ISTAT comune      Col 37  Tasso GB intero
  Col  2  zona AB comune           Col 38  Tasso GB Agevolato
  Col  3  codice CIAG comune       Col 39  Tasso GB Non agev
  Col  4  Descrizione comune       Col 40  Tasso GB_prim_aut intero
  Col  5  codice specie            Col 41  Tasso GB_prim_aut Agevolato
  Col  6  descrizione specie       Col 42  Tasso GB_prim_aut Non agev
  Col  7  Fr Minima GR             Col 43-45  SI irrigua  (int/ag/na)
  Col  8  Fr Minima VF             Col 46-48  SI NON irrigua (int/ag/na)
  Col  9  Fr Minima CAT            Col 49-51  AL (int/ag/na)
  Col 10-21  GR fr10/15/20/30      Col 52     ST Non agev fr 30%
  Col 22-33  VF fr10/15/20/30      Col 53-55  EN (int/ag/na)
  Col 34-36  EP fr30               Col 56-58  CS-VC (int/ag/na)
                                    Col 59-61  Accessorie (int/ag/na)
                                    Col 62     Cod_Tipologia_Prodotto

"Sconti" ha inoltre:  Col 64 Raggruppamento specie, Col 65 PERC VF AGEV, Col 66 RAGGR. ALTRE AVV.
"Tariffa_Normale" termina a Col 62.

Vengono scritti come:
  versione_listino = 1  ->  Tariffa Normale
  versione_listino = 2  ->  Sconti

Deduplicazione: per duplicati su (codice ISTAT, codice specie)
si tiene la riga con la somma dei tassi piu bassa.
"""
from __future__ import annotations

from typing import IO

import openpyxl
from sqlalchemy.orm import Session

from app.models import Tariffa
from app.calcolo import TIPO_GARANZIA

# Codice versione per distinguere i due sheet nel DB
VERSIONE_TARIFFA_NORMALE = 1
VERSIONE_SCONTI = 2


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


# Mappatura garanzie con multi-franchigia: (garanzia, [(fr%, int_col, ag_col, na_col), ...])
GARANZIE_MULTI_FR = {
    "grandine": [
        (10, 10, 11, 12),
        (15, 13, 14, 15),
        (20, 16, 17, 18),
        (30, 19, 20, 21),
    ],
    "vento_forte": [
        (10, 22, 23, 24),
        (15, 25, 26, 27),
        (20, 28, 29, 30),
        (30, 31, 32, 33),
    ],
    "eccesso_pioggia": [
        (30, 34, 35, 36),
    ],
}

# Garanzie con franchigia unica (catastrofali e accessorie)
GARANZIE_SINGOLE = {
    "gelo_brina":             (37, 38, 39),
    "siccita":                (43, 44, 45),
    "alluvione":              (49, 50, 51),
    "eccesso_neve":           (53, 54, 55),
    "colpo_sole_vento_caldo": (56, 57, 58),
    "sbalzo_termico":         (None, None, 52),
}


def _parse_sheet(ws, db, anno, versione, has_raggruppamento_col):
    """Parsa un singolo sheet RealeMutua. Restituisce n. record inseriti."""
    raggrupp_col = 64 if has_raggruppamento_col else None

    # Prima passata: raccogli dati e somma tassi per deduplicazione
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        get = lambda c, _r=row_idx: ws.cell(_r, c).value

        istat_raw = _safe_str(get(1))
        specie_raw = _safe_str(get(5))
        if not istat_raw and not specie_raw:
            continue

        tasso_sum = 0.0
        for c in range(10, 62):
            tasso_sum += abs(_safe_float(get(c)))

        rows.append({
            "row_idx": row_idx,
            "istat": istat_raw,
            "ciag": _safe_str(get(3)),
            "comune_nome": _safe_str(get(4)),
            "specie_cod": specie_raw,
            "specie_desc": _safe_str(get(6)),
            "fr_gr": _safe_float(get(7)),
            "fr_vf": _safe_float(get(8)),
            "fr_cat": _safe_float(get(9)),
            "raggruppamento": _safe_str(get(raggrupp_col)) if raggrupp_col else "",
            "tasso_sum": tasso_sum,
        })

    # Deduplicazione: per (ISTAT, codice specie) tieni la riga con tassi piu bassi
    best = {}
    for r in rows:
        key = (r["istat"], r["specie_cod"])
        if key not in best or r["tasso_sum"] < best[key]["tasso_sum"]:
            best[key] = r

    # Seconda passata: inserisci le righe selezionate
    count = 0
    for r in best.values():
        row_idx = r["row_idx"]
        get = lambda c, _r=row_idx: ws.cell(_r, c).value

        istat = r["istat"]
        if istat and istat.replace(".", "").isdigit():
            istat = str(int(float(istat))).zfill(6)

        # Garanzie multi-franchigia (GR, VF, EP)
        for garanzia, fr_list in GARANZIE_MULTI_FR.items():
            fr_min = r["fr_gr"]
            if garanzia == "vento_forte":
                fr_min = r["fr_vf"]

            for fr_pct, int_col, ag_col, na_col in fr_list:
                tasso_int = _safe_float(get(int_col))
                tasso_ag  = _safe_float(get(ag_col))
                tasso_na  = _safe_float(get(na_col))

                if abs(tasso_int) < 1e-9 and abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                    continue

                if abs(tasso_int) < 1e-9:
                    tasso_int = round(tasso_ag + tasso_na, 4)

                db.add(Tariffa(
                    compagnia="RealeMutua",
                    provincia="",
                    comune_istat=istat,
                    comune_ciag=r["ciag"],
                    comune_nome=r["comune_nome"],
                    specie_codice=r["specie_cod"],
                    specie_descrizione=r["specie_desc"],
                    raggruppamento=r["raggruppamento"],
                    garanzia=garanzia,
                    tipo_garanzia=TIPO_GARANZIA.get(garanzia, "frequenziale"),
                    franchigia_min=fr_min,
                    franchigia_applicata=float(fr_pct),
                    tasso_agevolato=tasso_ag,
                    tasso_non_agevolato=tasso_na,
                    tasso_totale=tasso_int,
                    anno_validita=anno,
                    versione_listino=versione,
                ))
                count += 1

        # Garanzie singole (catastrofali)
        fr_cat = r["fr_cat"]
        for garanzia, (int_col, ag_col, na_col) in GARANZIE_SINGOLE.items():
            tasso_int = _safe_float(get(int_col)) if int_col else 0.0
            tasso_ag  = _safe_float(get(ag_col))  if ag_col  else 0.0
            tasso_na  = _safe_float(get(na_col))  if na_col  else 0.0

            if abs(tasso_int) < 1e-9 and abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                continue

            if abs(tasso_int) < 1e-9:
                tasso_int = round(tasso_ag + tasso_na, 4)
            if abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                tasso_ag = tasso_int

            db.add(Tariffa(
                compagnia="RealeMutua",
                provincia="",
                comune_istat=istat,
                comune_ciag=r["ciag"],
                comune_nome=r["comune_nome"],
                specie_codice=r["specie_cod"],
                specie_descrizione=r["specie_desc"],
                raggruppamento=r["raggruppamento"],
                garanzia=garanzia,
                tipo_garanzia=TIPO_GARANZIA.get(garanzia, "catastrofale"),
                franchigia_min=fr_cat,
                franchigia_applicata=fr_cat if fr_cat > 0 else 30.0,
                tasso_agevolato=tasso_ag,
                tasso_non_agevolato=tasso_na,
                tasso_totale=tasso_int,
                anno_validita=anno,
                versione_listino=versione,
            ))
            count += 1

    return count


def parse_reale_mutua(file, db, anno=2026, versione=1):
    """Parsa entrambi gli sheet di RealeMutua.xlsx. Restituisce record totali."""
    wb = openpyxl.load_workbook(file, data_only=True)

    total = 0

    # Sheet "Tariffa_Normale" -> versione_listino = 1
    if "Tariffa_Normale" in wb.sheetnames:
        ws_normale = wb["Tariffa_Normale"]
        total += _parse_sheet(ws_normale, db, anno, VERSIONE_TARIFFA_NORMALE,
                              has_raggruppamento_col=False)
        db.flush()

    # Sheet "Sconti" -> versione_listino = 2
    if "Sconti" in wb.sheetnames:
        ws_sconti = wb["Sconti"]
        total += _parse_sheet(ws_sconti, db, anno, VERSIONE_SCONTI,
                              has_raggruppamento_col=True)

    wb.close()
    db.commit()
    return total
