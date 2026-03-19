"""
Parser per REVO.xlsx
Sheet "tutti i prodotti": Garanzie Frequenziali per Provincia/Comune/Specie
  Riga 4 = header; Riga 5+ = dati
  Col layout (1-based):
    1=Provincia, 2=Codice Comune, 3=Comune, 4=Raggruppamento,
    5=Codice Specie, 6=Specie,
    7=GRANDINE(label), 8=Fr Min GR, 9=Tasso Ag GR, 10=Tasso NA GR,
    11=VENTO FORTE(label), 12=Fr Min VF, 13=Tasso Ag VF, 14=Tasso NA VF,
    15=ECCESSO DI PIOGGIA(label), 16=Fr Min EP, 17=Tasso Ag EP, 18=Tasso NA EP

Sheet "altre garanzie": Catastrofali + extra per Raggruppamento
  Riga 1 = header; Riga 2+ = dati
  Col: 1=Raggruppamento, 2=GARANZIA, 3=Franchigia Min, 4=Tasso Ag, 5=Tasso NA
"""
from __future__ import annotations

from typing import IO

import openpyxl
from sqlalchemy.orm import Session

from app.models import Tariffa
from app.calcolo import TIPO_GARANZIA, normalizza_garanzia


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


# Blocchi frequenziali: (garanzia_norm, fr_col, ag_col, na_col)
FREQ_BLOCKS = [
    ("grandine",        8,  9, 10),
    ("vento_forte",    12, 13, 14),
    ("eccesso_pioggia",16, 17, 18),
]


def _parse_tutti_i_prodotti(wb, db, anno, versione):
    """Sheet 'tutti i prodotti': frequenziali per Provincia/Comune/Specie."""
    ws = wb["tutti i prodotti"]
    count = 0

    for row_idx in range(5, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        cod_comune = _safe_str(get(2))
        if not cod_comune:
            continue

        provincia      = _safe_str(get(1))
        comune_nome    = _safe_str(get(3))
        raggruppamento = _safe_str(get(4))
        specie_cod     = _safe_str(get(5))
        specie_desc    = _safe_str(get(6))

        for garanzia, fr_col, ag_col, na_col in FREQ_BLOCKS:
            fr = _safe_float(get(fr_col))
            ag = _safe_float(get(ag_col))
            na = _safe_float(get(na_col))

            if abs(ag) < 1e-9 and abs(na) < 1e-9:
                continue

            db.add(Tariffa(
                compagnia="REVO",
                provincia=provincia,
                comune_istat="",
                comune_ciag=cod_comune,
                comune_nome=comune_nome,
                specie_codice=specie_cod,
                specie_descrizione=specie_desc,
                raggruppamento=raggruppamento,
                garanzia=garanzia,
                tipo_garanzia="frequenziale",
                franchigia_min=fr,
                franchigia_applicata=fr,
                tasso_agevolato=ag,
                tasso_non_agevolato=na,
                tasso_totale=round(ag + na, 4),
                anno_validita=anno,
                versione_listino=versione,
            ))
            count += 1

    return count


def _parse_altre_garanzie(wb, db, anno, versione):
    """Sheet 'altre garanzie': catastrofali + extra per raggruppamento."""
    ws = wb["altre garanzie"]
    count = 0

    for row_idx in range(2, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        raggruppamento = _safe_str(get(1))
        garanzia_raw   = _safe_str(get(2))
        if not garanzia_raw:
            continue

        garanzia_norm = normalizza_garanzia(garanzia_raw, "revo")
        if not garanzia_norm or garanzia_norm not in TIPO_GARANZIA:
            # Es. "TAB D EXTRA Q", "cod 005A" -> non mappabili, skip
            continue

        fr = _safe_float(get(3))
        ag = _safe_float(get(4))
        na = _safe_float(get(5))

        if abs(ag) < 1e-9 and abs(na) < 1e-9:
            continue

        db.add(Tariffa(
            compagnia="REVO",
            provincia="",
            comune_istat="",
            comune_ciag="",
            comune_nome="",
            specie_codice="",
            specie_descrizione="",
            raggruppamento=raggruppamento,
            garanzia=garanzia_norm,
            tipo_garanzia=TIPO_GARANZIA.get(garanzia_norm, "catastrofale"),
            franchigia_min=fr,
            franchigia_applicata=fr,
            tasso_agevolato=ag,
            tasso_non_agevolato=na,
            tasso_totale=round(ag + na, 4),
            anno_validita=anno,
            versione_listino=versione,
        ))
        count += 1

    return count


def parse_revo(file, db, anno=2026, versione=1):
    """Parsa entrambi gli sheet di REVO.xlsx. Restituisce record totali."""
    wb = openpyxl.load_workbook(file, data_only=True)
    c1 = _parse_tutti_i_prodotti(wb, db, anno, versione)
    c2 = _parse_altre_garanzie(wb, db, anno, versione)
    wb.close()
    db.commit()
    return c1 + c2
