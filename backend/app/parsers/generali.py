"""
Parser per Generali.xlsx — sheet "Tariffario".

Layout (anno 2026):
  Riga 1: intestazioni macro-area
  Riga 2: nomi colonna veri
  Riga 3+: dati
"""
from __future__ import annotations

from typing import IO

import openpyxl
from sqlalchemy.orm import Session

from app.models import Tariffa
from app.calcolo import TIPO_GARANZIA

# Mappatura posizionale 1-based -> (franchigia, agevolato, non_agev, totale)
GARANZIE_MAP = {
    "grandine":                (12, 13, 14, 15),
    "eccesso_pioggia":         (None, 16, 17, 18),
    "vento_forte":             (22, 23, 24, 25),
    "eccesso_neve":            (None, 26, 27, 28),
    "gelo_brina":              (29, None, None, 30),
    "siccita":                 (29, None, None, 31),
    "alluvione":               (29, None, None, 32),
    "colpo_sole_vento_caldo":  (None, 33, 34, 35),
    "sbalzo_termico":          (None, 36, 37, 38),
}

FRANCHIGIA_FISSA = {
    "eccesso_pioggia": 30.0,
    "eccesso_neve": 30.0,
    "colpo_sole_vento_caldo": 30.0,
    "sbalzo_termico": 30.0,
}


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        f = float(val)
        return f if f == f else 0.0
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def parse_generali(file, db, anno=2026, versione=1):
    """Parsa Generali.xlsx sheet Tariffario. Restituisce record inseriti."""
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb["Tariffario"]
    count = 0

    for row_idx in range(3, ws.max_row + 1):
        get = lambda c: ws.cell(row_idx, c).value

        comune_istat_raw = _safe_str(get(4))
        if not comune_istat_raw:
            continue

        comune_istat = comune_istat_raw.zfill(6) if comune_istat_raw.isdigit() else comune_istat_raw
        provincia      = _safe_str(get(2))
        comune_ciag    = _safe_str(get(3))
        comune_nome    = _safe_str(get(5))
        specie_cod     = _safe_str(get(6))
        specie_desc    = _safe_str(get(9))
        raggruppamento = _safe_str(get(11))
        fr_cat = _safe_float(get(29))

        for garanzia, (fr_col, ag_col, na_col, tot_col) in GARANZIE_MAP.items():
            tasso_ag  = _safe_float(get(ag_col))  if ag_col  else 0.0
            tasso_na  = _safe_float(get(na_col))  if na_col  else 0.0
            tasso_tot = _safe_float(get(tot_col)) if tot_col else 0.0

            if abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9 and abs(tasso_tot) < 1e-9:
                continue

            if abs(tasso_tot) < 1e-9:
                tasso_tot = round(tasso_ag + tasso_na, 4)

            if abs(tasso_ag) < 1e-9 and abs(tasso_na) < 1e-9:
                tasso_ag = tasso_tot

            if fr_col is not None:
                franchigia = _safe_float(get(fr_col))
            else:
                franchigia = FRANCHIGIA_FISSA.get(garanzia, fr_cat)

            db.add(Tariffa(
                compagnia="Generali",
                provincia=provincia,
                comune_istat=comune_istat,
                comune_ciag=comune_ciag,
                comune_nome=comune_nome,
                specie_codice=specie_cod,
                specie_descrizione=specie_desc,
                raggruppamento=raggruppamento,
                garanzia=garanzia,
                tipo_garanzia=TIPO_GARANZIA.get(garanzia, "frequenziale"),
                franchigia_min=franchigia,
                franchigia_applicata=franchigia,
                tasso_agevolato=tasso_ag,
                tasso_non_agevolato=tasso_na,
                tasso_totale=tasso_tot,
                anno_validita=anno,
                versione_listino=versione,
            ))
            count += 1

            if count % 5000 == 0:
                db.flush()

    wb.close()
    db.commit()
    return count
